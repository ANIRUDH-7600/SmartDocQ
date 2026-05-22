from __future__ import annotations

import csv
import io
import re
import uuid
from dataclasses import dataclass
from hashlib import sha1
from typing import Any, Iterable, Iterator, Optional, Sequence, TypedDict


# =========================
# Public schema
# =========================


class ExtractedTable(TypedDict):
    table_id: str
    sheet: str | None
    headers: list[str]
    rows: list[list[str]]
    source_page: None
    markdown: str
    flattened_text: str
    raw_text: str


# =========================
# Normalization helpers
# =========================


_EMPTY_RE = re.compile(r"^\s*$")
_NUMERIC_RE = re.compile(r"^\s*[-+]?\d+(?:\.\d+)?\s*$")


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    # openpyxl may return numbers, datetimes, booleans
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


def _is_row_empty(row: Sequence[str]) -> bool:
    return all(_EMPTY_RE.match(c or "") for c in row)


def _pad_row(row: list[str], width: int) -> list[str]:
    if len(row) >= width:
        return row[:width]
    return row + [""] * (width - len(row))


def _looks_like_header(row: Sequence[str]) -> bool:
    cells = [c.strip() for c in row if (c or "").strip()]
    if not cells:
        return False

    # If mostly numeric, it's probably not a header
    numeric = sum(1 for c in cells if _NUMERIC_RE.match(c))
    if numeric / max(len(cells), 1) >= 0.6:
        return False

    # Prefer rows with low duplication
    uniq = len({c.lower() for c in cells})
    if uniq / max(len(cells), 1) < 0.7:
        return False

    # Require at least one alpha character somewhere
    if not any(any(ch.isalpha() for ch in c) for c in cells):
        return False

    return True


def render_markdown_table(headers: Sequence[str], rows: Sequence[Sequence[str]]) -> str:
    """Render a GitHub-flavored Markdown table."""

    hdr = [h.strip() for h in (headers or [])]
    width = max(len(hdr), max((len(r) for r in rows), default=0))

    if width == 0:
        return ""

    if not hdr:
        hdr = [f"Column {i + 1}" for i in range(width)]
    hdr = _pad_row(list(hdr), width)

    def fmt_row(r: Sequence[str]) -> str:
        rr = _pad_row([str(c).replace("\n", " ").strip() for c in r], width)
        # Keep pipes safe
        rr = [c.replace("|", "\\|") for c in rr]
        return "| " + " | ".join(rr) + " |"

    sep = "| " + " | ".join(["-----"] * width) + " |"

    out = [fmt_row(hdr), sep]
    for r in rows:
        out.append(fmt_row(r))
    return "\n".join(out)


def flatten_table_for_embedding(
    *,
    sheet: str | None,
    headers: Sequence[str],
    rows: Sequence[Sequence[str]],
) -> str:
    """Produce a semantic flattened text representation optimized for embeddings."""

    width = max(len(headers), max((len(r) for r in rows), default=0))
    hdr = _pad_row([h.strip() for h in headers], width)

    lines: list[str] = []
    if sheet:
        lines.append(f"Sheet: {sheet}")
        lines.append("")

    for row in rows:
        rr = _pad_row([str(c).strip() for c in row], width)
        if all(not c for c in rr):
            continue

        lines.append("Row:")
        for h, v in zip(hdr, rr):
            key = (h or "Column").strip() or "Column"
            lines.append(f"{key} = {v}")
        lines.append("")

    return "\n".join(lines).strip()


def _raw_text_from_table(headers: Sequence[str], rows: Sequence[Sequence[str]], sheet: str | None) -> str:
    width = max(len(headers), max((len(r) for r in rows), default=0))
    hdr = _pad_row([h.strip() for h in headers], width)

    parts: list[str] = []
    if sheet:
        parts.append(f"# Sheet: {sheet}")

    parts.append("\t".join(hdr))
    for r in rows:
        rr = _pad_row([str(c).strip() for c in r], width)
        if _is_row_empty(rr):
            continue
        parts.append("\t".join(rr))

    return "\n".join(parts).strip()


def _make_table_id(source_key: str | None, sheet: str | None, table_index: int, headers: list[str], rows: list[list[str]]) -> str:
    """Deterministic-ish ID when source_key is provided; otherwise random."""

    if not source_key:
        return uuid.uuid4().hex

    sample_rows = rows[:5]
    payload = {
        "source": source_key,
        "sheet": sheet or "",
        "table_index": table_index,
        "headers": headers,
        "rows": sample_rows,
    }
    digest = sha1(repr(payload).encode("utf-8", errors="ignore")).hexdigest()
    return digest


# =========================
# Extractors
# =========================


def extract_tables_from_csv_bytes(
    data: bytes,
    *,
    sheet_name: str | None = "Sheet1",
    source_key: str | None = None,
    encoding: str = "utf-8",
) -> list[ExtractedTable]:
    """Extract a single table from CSV bytes.

    - Detects header row automatically (best-effort).
    - Preserves row order.
    """

    if not data:
        return []

    text = data.decode(encoding, errors="ignore")
    buf = io.StringIO(text)

    reader = csv.reader(buf)

    rows_raw: list[list[str]] = [[_cell_to_str(c) for c in r] for r in reader]
    # Drop fully empty rows
    rows_raw = [r for r in rows_raw if not _is_row_empty(r)]
    if not rows_raw:
        return []

    first = rows_raw[0]
    if _looks_like_header(first):
        headers = first
        data_rows = rows_raw[1:]
    else:
        width = max(len(r) for r in rows_raw)
        headers = [f"Column {i + 1}" for i in range(width)]
        data_rows = rows_raw

    width = max(len(headers), max((len(r) for r in data_rows), default=0))
    headers = _pad_row([h.strip() for h in headers], width)
    rows_out = [_pad_row([c.strip() for c in r], width) for r in data_rows if not _is_row_empty(r)]

    md = render_markdown_table(headers, rows_out)
    flat = flatten_table_for_embedding(sheet=sheet_name, headers=headers, rows=rows_out)
    raw = _raw_text_from_table(headers, rows_out, sheet_name)

    table_id = _make_table_id(source_key, sheet_name, 0, headers, rows_out)

    return [
        {
            "table_id": table_id,
            "sheet": sheet_name,
            "headers": list(headers),
            "rows": [list(r) for r in rows_out],
            "source_page": None,
            "markdown": md,
            "flattened_text": flat,
            "raw_text": raw,
        }
    ]


def _iter_xlsx_rows(values_iter: Iterable[Sequence[Any]]) -> Iterator[list[str]]:
    for row in values_iter:
        yield [_cell_to_str(v) for v in (row or [])]


def _split_tables_by_blank_rows(rows: list[list[str]]) -> list[list[list[str]]]:
    """Split a sheet into table blocks separated by fully empty rows."""

    blocks: list[list[list[str]]] = []
    cur: list[list[str]] = []

    for r in rows:
        if _is_row_empty(r):
            if cur:
                blocks.append(cur)
                cur = []
            continue
        cur.append(r)

    if cur:
        blocks.append(cur)

    return blocks


def extract_tables_from_xlsx_bytes(
    data: bytes,
    *,
    source_key: str | None = None,
    max_rows_per_table: int | None = None,
) -> list[ExtractedTable]:
    """Extract tables from XLSX bytes.

    - Uses openpyxl in read-only mode (streaming-ish).
    - Supports multiple sheets.
    - Splits multiple tables per sheet by blank rows.
    """

    if not data:
        return []

    try:
        from openpyxl import load_workbook
    except Exception as e:  # pragma: no cover
        # Fail-soft: treat missing optional dependency as "no tables".
        return []

    tables: list[ExtractedTable] = []

    try:
        with io.BytesIO(data) as bio:
            wb = load_workbook(bio, read_only=True, data_only=True)

            for sheet in wb.worksheets:
                sheet_name = sheet.title

                # Streaming row iteration.
                sheet_rows: list[list[str]] = []
                n = 0
                for r in _iter_xlsx_rows(sheet.iter_rows(values_only=True)):
                    sheet_rows.append(r)
                    n += 1
                    if max_rows_per_table and n >= int(max_rows_per_table):
                        # In phase 1, we keep this simple: truncate extremely large sheets.
                        break

                if not sheet_rows:
                    continue

                blocks = _split_tables_by_blank_rows(sheet_rows)
                table_idx = 0
                for block in blocks:
                    block = [r for r in block if not _is_row_empty(r)]
                    if not block:
                        continue

                    first = block[0]
                    if _looks_like_header(first):
                        headers = first
                        data_rows = block[1:]
                    else:
                        width = max(len(r) for r in block)
                        headers = [f"Column {i + 1}" for i in range(width)]
                        data_rows = block

                    width = max(len(headers), max((len(r) for r in data_rows), default=0))
                    headers = _pad_row([h.strip() for h in headers], width)
                    rows_out = [_pad_row([c.strip() for c in r], width) for r in data_rows if not _is_row_empty(r)]

                    # Ignore tables with no body.
                    if not rows_out:
                        continue

                    md = render_markdown_table(headers, rows_out)
                    flat = flatten_table_for_embedding(sheet=sheet_name, headers=headers, rows=rows_out)
                    raw = _raw_text_from_table(headers, rows_out, sheet_name)

                    table_id = _make_table_id(source_key, sheet_name, table_idx, headers, rows_out)

                    tables.append(
                        {
                            "table_id": table_id,
                            "sheet": sheet_name,
                            "headers": list(headers),
                            "rows": [list(r) for r in rows_out],
                            "source_page": None,
                            "markdown": md,
                            "flattened_text": flat,
                            "raw_text": raw,
                        }
                    )

                    table_idx += 1
    except Exception:
        # Fail-soft on corrupted / partially uploaded XLSX files.
        return []

    return tables


def extract_tables_from_docx_bytes(
    data: bytes,
    *,
    source_key: str | None = None,
    sheet_name: str | None = None,
) -> list[ExtractedTable]:
    """Extract all tables from a DOCX.

    - Preserves table order.
    - Treats each Word table as a single extracted table.
    """

    if not data:
        return []

    try:
        from docx import Document as DocxDocument
    except Exception as e:  # pragma: no cover
        raise RuntimeError("python-docx is required for DOCX table extraction") from e

    out: list[ExtractedTable] = []

    with io.BytesIO(data) as f:
        doc = DocxDocument(f)

        table_index = 0
        for t in doc.tables:
            block: list[list[str]] = []
            for row in t.rows:
                block.append([_cell_to_str(c.text) for c in row.cells])

            block = [r for r in block if not _is_row_empty(r)]
            if not block:
                continue

            first = block[0]
            if _looks_like_header(first):
                headers = first
                data_rows = block[1:]
            else:
                width = max(len(r) for r in block)
                headers = [f"Column {i + 1}" for i in range(width)]
                data_rows = block

            width = max(len(headers), max((len(r) for r in data_rows), default=0))
            headers = _pad_row([h.strip() for h in headers], width)
            rows_out = [_pad_row([c.strip() for c in r], width) for r in data_rows if not _is_row_empty(r)]

            if not rows_out:
                continue

            md = render_markdown_table(headers, rows_out)
            flat = flatten_table_for_embedding(sheet=sheet_name, headers=headers, rows=rows_out)
            raw = _raw_text_from_table(headers, rows_out, sheet_name)

            table_id = _make_table_id(source_key, sheet_name, table_index, headers, rows_out)

            out.append(
                {
                    "table_id": table_id,
                    "sheet": sheet_name,
                    "headers": list(headers),
                    "rows": [list(r) for r in rows_out],
                    "source_page": None,
                    "markdown": md,
                    "flattened_text": flat,
                    "raw_text": raw,
                }
            )

            table_index += 1

    return out


def extract_tables_for_file(filename: str, mimetype: str, data: bytes, *, source_key: str | None = None) -> list[ExtractedTable]:
    ext = (filename.rsplit(".", 1)[-1].lower() if "." in (filename or "") else "")
    mt = (mimetype or "").lower()

    if mt == "text/csv" or ext == "csv":
        return extract_tables_from_csv_bytes(data, sheet_name="Sheet1", source_key=source_key)

    if mt == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" or ext in ("xlsx",):
        return extract_tables_from_xlsx_bytes(data, source_key=source_key)

    if mt in (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    ) or ext in ("docx", "doc"):
        # For Word tables, we keep sheet None to avoid implying pages/sheets.
        return extract_tables_from_docx_bytes(data, source_key=source_key, sheet_name=None)

    return []


def tables_to_scan_text(tables: Sequence[ExtractedTable]) -> str:
    """Build a text blob representing extracted tables for sensitive-data scanning."""

    parts: list[str] = []
    for t in tables:
        # Prefer raw text (compact) + a flattened view (semantic)
        raw = (t.get("raw_text") or "").strip()
        flat = (t.get("flattened_text") or "").strip()
        if raw:
            parts.append(raw)
        if flat and flat != raw:
            parts.append(flat)

    return "\n\n".join(parts).strip()
